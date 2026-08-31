# -*- coding: utf-8 -*-
"""チラシ進行台帳（Excel版）を組み立てる"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

OUT = "/home/user/test/excel/チラシ進行台帳.xlsx"
FONT = "Meiryo"
MAX_S = 71           # 学校マスタ 最終行（70校）
MAX_P = 251          # 価格 最終行（250行）
WEEKS = 24

INK   = "1F2430"
MUTED = "6B7280"
LINE  = "C9D2E3"
HEAD  = "1F3864"      # 見出し背景
HEADT = "FFFFFF"
BAND  = "F2F5FB"
BLUE  = "0000FF"      # 手入力
GREEN = "006400"      # 他シート参照
BLACK = "000000"      # 計算
YEL   = "FFF7D6"

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row, ncol, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color=HEADT)
        cell.fill = PatternFill("solid", fgColor=HEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- 使い方
ws = wb.active
ws.title = "使い方"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 96})
rows = [
    ("h1", "チラシ進行台帳（Excel版）", ""),
    ("p", "", "学校ごとの工程管理と、価格・手数料バックの管理を1つのブックで行います。シート間は数式でつながっているので、"
            "「学校マスタ」に学校を足し、「価格」に商品を足すだけで、ガントも手数料集計も自動で追従します。"),
    ("h2", "シートの役割", ""),
    ("li", "学校マスタ", "学校ごとに1行。工程のチェックと期日、協力企業、手数料バックの有無。ここが台帳の本体です。"),
    ("li", "価格", "学校×商品ごとに1行。PC機種もオプション品もここに登録します。手数料バックはこのシートで自動計算されます。"),
    ("li", "ガント", "学校マスタの日付から自動で帯を描きます。入力する場所はありません（開始日だけ変更できます）。"),
    ("li", "手数料集計", "協力企業ごとに、戻す手数料の見込みを集計します。全体のサマリーもここにあります。"),
    ("li", "設定", "担当・地区・協力企業・納品方法の選択肢。ここに足すと各シートのプルダウンに出ます。"),
    ("h2", "入力する場所（セルの色）", ""),
    ("li", "青い文字", "手で入力する欄です。ここだけ触ってください。"),
    ("li", "黒い文字", "同じシート内の計算結果。上書きすると壊れます。"),
    ("li", "緑の文字", "他のシートを参照している計算結果。上書きすると壊れます。"),
    ("li", "黄色の背景", "最初に決めておく設定値（ガントの開始日など）。"),
    ("h2", "使い方の順番", ""),
    ("li", "1. 設定", "担当者名・地区・協力企業・納品方法を先に登録します。"),
    ("li", "2. 学校マスタ", "学校名から順に入力。日付を入れると状態とガントが自動で動きます。工程の「済」列はプルダウンで ✓ を選びます。"),
    ("li", "3. 価格", "学校名をプルダウンで選び、種別（PC／オプション）と品名、今年の掲載・卸・見込数量を入れます。"),
    ("li", "4. 手数料集計", "自動で集計されます。数字が合わないときは、価格シートの学校名が学校マスタと一致しているか確認してください。"),
    ("h2", "手数料バックの計算", ""),
    ("p", "", "手数料バック単価 ＝ 今年 掲載（エンド税別） − 今年 卸（税別）。"
            "学校マスタの「手数料バック」を「なし」にすると、その学校の手数料は 0 になり、集計にも入りません。"
            "見込合計 ＝ 単価 × 見込 数量 です。"),
    ("h2", "注意", ""),
    ("li", "学校名がキー", "価格シートは学校名で学校マスタと結び付いています。学校名を変えるときは両方のシートを直してください（重複した学校名は使えません）。"),
    ("li", "行の追加", "数式は70校・250商品ぶん入れてあります。それ以上増やすときは、最終行をコピーして下に貼り付けてください。"),
    ("li", "サンプル行", "1校ぶんのサンプルが入っています。実データを入れる前に削除してください（数式の列は消さずに、青文字の欄だけ消します）。"),
]
r = 1
for kind, label, text in rows:
    if kind == "h1":
        c = ws.cell(row=r, column=2, value=label)
        c.font = Font(name=FONT, bold=True, size=16, color=HEAD)
        ws.row_dimensions[r].height = 26
    elif kind == "h2":
        r += 1
        c = ws.cell(row=r, column=2, value=label)
        c.font = Font(name=FONT, bold=True, size=11, color=HEAD)
    elif kind == "li":
        c = ws.cell(row=r, column=2, value=label)
        c.font = Font(name=FONT, bold=True, size=10, color=INK)
        c.alignment = Alignment(vertical="top")
        t = ws.cell(row=r, column=3, value=text)
        t.font = Font(name=FONT, size=10, color=MUTED)
        t.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
    else:
        t = ws.cell(row=r, column=3, value=text)
        t.font = Font(name=FONT, size=10, color=MUTED)
        t.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 46
    r += 1

# ---------------------------------------------------------------- 設定
st = wb.create_sheet("設定")
st.sheet_view.showGridLines = False
setting_cols = [
    ("担当", ["内野", "佐藤", "高橋", "森田", "岡本"]),
    ("地区", ["東京", "神奈川", "千葉", "埼玉", "大阪", "愛知", "福岡", "北海道", "宮城", "広島", "京都", "兵庫"]),
    ("協力企業", ["サクラ電機販売", "三洋メディアサービス", "日商テクノ商会", "アイエス情報機器", "（協力企業なし）"]),
    ("納品方法", ["一括納品（学校搬入）", "個別配送（自宅）", "説明会当日手渡し", "業者直送（分割）"]),
    ("区分", ["継続", "新規開拓"]),
    ("種別", ["PC", "オプション"]),
    ("手数料バック", ["あり", "なし"]),
]
for i, (name, vals) in enumerate(setting_cols, start=1):
    st.cell(row=1, column=i, value=name)
    for j, v in enumerate(vals, start=2):
        c = st.cell(row=j, column=i, value=v)
        c.font = Font(name=FONT, size=10, color=BLUE)
        c.border = box
style_header(st, 1, len(setting_cols))
widths(st, {L(i): 22 for i in range(1, len(setting_cols) + 1)})
st.cell(row=20, column=1, value="※ 各列に追記すると、他シートのプルダウンの選択肢が増えます（20行目まで対応）。").font = Font(name=FONT, size=9, color=MUTED)

# ---------------------------------------------------------------- 学校マスタ
ms = wb.create_sheet("学校マスタ")
M_HEAD = [
    "学校名", "地区", "担当", "区分", "生徒数", "協力企業", "手数料バック",
    "機種提示日", "済", "意向回収期日", "済", "EC公開期日", "済", "ECサイトURL",
    "チラシ調整開始", "チラシ調整終了", "済", "販売開始日", "済",
    "納品方法", "納品期日", "済",
    "進捗", "状態", "掲載機種数", "オプション点数", "価格未入力", "手数料バック見込", "メモ",
]
C = {h: L(i) for i, h in enumerate(M_HEAD, start=1)}   # 重複見出し「済」は最後の列になるため個別に持つ
DONE = {"model": "I", "intent": "K", "ec": "M", "flyer": "Q", "sale": "S", "deliv": "V"}
DATE = {"model": "H", "intent": "J", "ec": "L", "flyerS": "O", "flyerE": "P", "sale": "R", "deliv": "U"}
for i, h in enumerate(M_HEAD, start=1):
    ms.cell(row=1, column=i, value=h)
style_header(ms, 1, len(M_HEAD), height=30)
widths(ms, {"A": 26, "B": 10, "C": 9, "D": 10, "E": 8, "F": 20, "G": 12,
            "H": 12, "I": 5, "J": 12, "K": 5, "L": 12, "M": 5, "N": 26,
            "O": 13, "P": 13, "Q": 5, "R": 12, "S": 5,
            "T": 20, "U": 12, "V": 5,
            "W": 8, "X": 10, "Y": 10, "Z": 12, "AA": 10, "AB": 16, "AC": 30})
ms.freeze_panes = "B2"

input_cols = list("ABCDEFGHIJKLMNOPQRSTUV") + ["AC"]
date_cols = ["H", "J", "L", "O", "P", "R", "U"]
for r in range(2, MAX_S + 1):
    done_terms = " ".join([f'COUNTIF({d}{r},"✓")+' for d in DONE.values()]).rstrip("+")
    late = ",".join([f'AND({DONE[k]}{r}<>"✓",{DATE[dk]}{r}<>"",{DATE[dk]}{r}<TODAY())'
                     for k, dk in [("model","model"),("intent","intent"),("ec","ec"),
                                   ("flyer","flyerE"),("sale","sale"),("deliv","deliv")]])
    soon = ",".join([f'AND({DONE[k]}{r}<>"✓",{DATE[dk]}{r}<>"",{DATE[dk]}{r}<=TODAY()+7)'
                     for k, dk in [("model","model"),("intent","intent"),("ec","ec"),
                                   ("flyer","flyerE"),("sale","sale"),("deliv","deliv")]])
    ms[f"W{r}"] = f'=IF($A{r}="","",({done_terms})/6)'
    ms[f"X{r}"] = (f'=IF($A{r}="","",IF(W{r}=1,"完了",'
                   f'IF(OR({late}),"遅延",IF(OR({soon}),"今週対応","進行中"))))')
    PA, PB, PH, PJ, PN = (f"価格!$A$2:$A${MAX_P}", f"価格!$B$2:$B${MAX_P}",
                          f"価格!$H$2:$H${MAX_P}", f"価格!$J$2:$J${MAX_P}", f"価格!$N$2:$N${MAX_P}")
    ms[f"Y{r}"] = f'=IF($A{r}="","",COUNTIFS({PA},$A{r},{PB},"PC"))'
    ms[f"Z{r}"] = f'=IF($A{r}="","",COUNTIFS({PA},$A{r},{PB},"オプション"))'
    ms[f"AA{r}"] = (f'=IF($A{r}="","",COUNTIFS({PA},$A{r})'
                    f'-COUNTIFS({PA},$A{r},{PH},">0",{PJ},">0"))')
    ms[f"AB{r}"] = f'=IF($A{r}="","",SUMIFS({PN},{PA},$A{r}))'
    for col in input_cols:
        c = ms[f"{col}{r}"]
        c.font = Font(name=FONT, size=10, color=BLUE)
        c.border = box
        if col in date_cols:
            c.number_format = "yyyy/mm/dd"
        if col == "E":
            c.number_format = "#,##0"
        if col in DONE.values():
            c.alignment = Alignment(horizontal="center")
    for col in ["W", "X", "Y", "Z", "AA", "AB"]:
        c = ms[f"{col}{r}"]
        c.font = Font(name=FONT, size=10, color=GREEN if col in ("Y","Z","AA","AB") else BLACK)
        c.border = box
        c.alignment = Alignment(horizontal="center")
    ms[f"W{r}"].number_format = "0%"
    ms[f"AB{r}"].number_format = '¥#,##0;-¥#,##0;"-"'
    if r % 2 == 0:
        for col in M_HEAD:
            pass
ms.auto_filter.ref = f"A1:AC{MAX_S}"

# 入力規則
dv_sets = [
    ("C", "設定!$A$2:$A$20"), ("B", "設定!$B$2:$B$20"), ("F", "設定!$C$2:$C$20"),
    ("T", "設定!$D$2:$D$20"), ("D", "設定!$E$2:$E$20"), ("G", "設定!$G$2:$G$20"),
]
for col, src in dv_sets:
    dv = DataValidation(type="list", formula1=src, allow_blank=True)
    ms.add_data_validation(dv)
    dv.add(f"{col}2:{col}{MAX_S}")
dv_chk = DataValidation(type="list", formula1='"✓"', allow_blank=True)
ms.add_data_validation(dv_chk)
for col in DONE.values():
    dv_chk.add(f"{col}2:{col}{MAX_S}")

# 状態の色分け
rng = f"X2:X{MAX_S}"
ms.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"遅延"'],
    fill=PatternFill("solid", fgColor="F8D7D5"), font=Font(name=FONT, size=10, bold=True, color="9B2C1E")))
ms.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"今週対応"'],
    fill=PatternFill("solid", fgColor="FBEED2"), font=Font(name=FONT, size=10, bold=True, color="8A5A12")))
ms.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"完了"'],
    fill=PatternFill("solid", fgColor="DCEFE7"), font=Font(name=FONT, size=10, color="14664D")))
for col in DONE.values():
    ms.conditional_formatting.add(f"{col}2:{col}{MAX_S}", CellIsRule(operator="equal", formula=['"✓"'],
        fill=PatternFill("solid", fgColor="DCEFE7"), font=Font(name=FONT, size=10, bold=True, color="14664D")))
ms.conditional_formatting.add(f"AA2:AA{MAX_S}", CellIsRule(operator="greaterThan", formula=["0"],
    fill=PatternFill("solid", fgColor="FBEED2")))
# 期日が過ぎていて未完了の日付を赤く
for k, dk in [("model","model"),("intent","intent"),("ec","ec"),("flyer","flyerE"),("sale","sale"),("deliv","deliv")]:
    dcol, ccol = DATE[dk], DONE[k]
    ms.conditional_formatting.add(f"{dcol}2:{dcol}{MAX_S}", FormulaRule(
        formula=[f'AND($A2<>"",{dcol}2<>"",{dcol}2<TODAY(),{ccol}2<>"✓")'],
        font=Font(name=FONT, size=10, bold=True, color="9B2C1E")))

ms.cell(row=1, column=1).comment = Comment(
    "学校名がキーです。価格シート・ガントシートはこの名前で紐づいています。重複しない名前を入れてください。", "台帳")

# ---------------------------------------------------------------- 価格
ps = wb.create_sheet("価格")
P_HEAD = ["学校名", "種別", "品名",
          "昨年 掲載", "昨年 卸", "昨年 台数", "昨年 売上",
          "今年 掲載（エンド税別）", "前年比", "今年 卸（税別）",
          "手数料バック", "手数料 単価", "見込 数量", "手数料 見込合計", "手数料率", "協力企業"]
for i, h in enumerate(P_HEAD, start=1):
    ps.cell(row=1, column=i, value=h)
style_header(ps, 1, len(P_HEAD), height=32)
widths(ps, {"A": 26, "B": 11, "C": 30, "D": 12, "E": 12, "F": 10, "G": 14,
            "H": 17, "I": 9, "J": 15, "K": 12, "L": 13, "M": 10, "N": 16, "O": 9, "P": 20})
ps.freeze_panes = "C2"
p_input = ["A", "B", "C", "D", "E", "F", "H", "J", "M"]
for r in range(2, MAX_P + 1):
    ps[f"G{r}"] = f'=IFERROR($D{r}*$F{r},"")'
    ps[f"I{r}"] = f'=IFERROR($H{r}/$D{r}-1,"")'
    ps[f"K{r}"] = f'=IF($A{r}="","",IFERROR(INDEX(学校マスタ!$G$2:$G${MAX_S},MATCH($A{r},学校マスタ!$A$2:$A${MAX_S},0)),""))'
    ps[f"L{r}"] = f'=IF(OR($A{r}="",$H{r}="",$J{r}=""),"",IF($K{r}="なし",0,$H{r}-$J{r}))'
    ps[f"N{r}"] = f'=IFERROR($L{r}*$M{r},"")'
    ps[f"O{r}"] = f'=IFERROR($L{r}/$H{r},"")'
    ps[f"P{r}"] = f'=IF($A{r}="","",IFERROR(INDEX(学校マスタ!$F$2:$F${MAX_S},MATCH($A{r},学校マスタ!$A$2:$A${MAX_S},0)),""))'
    for col in p_input:
        c = ps[f"{col}{r}"]
        c.font = Font(name=FONT, size=10, color=BLUE)
        c.border = box
        if col in ("D", "E", "H", "J"):
            c.number_format = '¥#,##0;-¥#,##0;"-"'
        if col in ("F", "M"):
            c.number_format = '#,##0;-#,##0;"-"'
    for col in ["G", "I", "K", "L", "N", "O", "P"]:
        c = ps[f"{col}{r}"]
        c.font = Font(name=FONT, size=10, color=GREEN if col in ("K", "P") else BLACK)
        c.border = box
    for col in ("G", "L", "N"):
        ps[f"{col}{r}"].number_format = '¥#,##0;-¥#,##0;"-"'
    ps[f"I{r}"].number_format = '+0.0%;-0.0%;"-"'
    ps[f"O{r}"].number_format = '0.0%;-0.0%;"-"'
ps.auto_filter.ref = f"A1:P{MAX_P}"
dv_school = DataValidation(type="list", formula1=f"学校マスタ!$A$2:$A${MAX_S}", allow_blank=True)
ps.add_data_validation(dv_school); dv_school.add(f"A2:A{MAX_P}")
dv_kind = DataValidation(type="list", formula1="設定!$F$2:$F$20", allow_blank=True)
ps.add_data_validation(dv_kind); dv_kind.add(f"B2:B{MAX_P}")
ps.conditional_formatting.add(f"B2:B{MAX_P}", CellIsRule(operator="equal", formula=['"オプション"'],
    fill=PatternFill("solid", fgColor="DCEFED"), font=Font(name=FONT, size=10, color="0F5E58")))
ps.conditional_formatting.add(f"B2:B{MAX_P}", CellIsRule(operator="equal", formula=['"PC"'],
    fill=PatternFill("solid", fgColor="E3E9F7"), font=Font(name=FONT, size=10, color="2F4B8F")))
ps.conditional_formatting.add(f"H2:H{MAX_P}", FormulaRule(
    formula=[f'AND($A2<>"",$H2="")'], fill=PatternFill("solid", fgColor=YEL)))
ps.conditional_formatting.add(f"J2:J{MAX_P}", FormulaRule(
    formula=[f'AND($A2<>"",$J2="")'], fill=PatternFill("solid", fgColor=YEL)))
ps.cell(row=1, column=12).comment = Comment(
    "手数料 単価 ＝ 今年 掲載（エンド税別） − 今年 卸（税別）。学校マスタで手数料バックを「なし」にした学校は 0 になります。", "台帳")

# ---------------------------------------------------------------- ガント
gs = wb.create_sheet("ガント")
gs.sheet_view.showGridLines = False
gs["A1"] = "ガント（1列＝1週間）"
gs["A1"].font = Font(name=FONT, bold=True, size=13, color=HEAD)
gs["E1"] = "█ チラシ調整期間　　▒ 販売開始→納品　　◆ 機種提示・意向回収・EC公開　　◎ 納品期日　　（今週の列は色が付きます）"
gs["E1"].font = Font(name=FONT, size=9, color=MUTED)
gs["A2"] = "開始日"
gs["A2"].font = Font(name=FONT, bold=True, size=10, color=INK)
gs["B2"] = dt.date(2026, 7, 6)
gs["B2"].font = Font(name=FONT, bold=True, size=10, color=BLUE)
gs["B2"].fill = PatternFill("solid", fgColor=YEL)
gs["B2"].number_format = "yyyy/mm/dd"
gs["B2"].border = box
gs["C2"] = "←月曜日を入れてください"
gs["C2"].font = Font(name=FONT, size=9, color=MUTED)

G_HEAD_ROW = 3
gs.cell(row=G_HEAD_ROW, column=1, value="学校名")
gs.cell(row=G_HEAD_ROW, column=2, value="担当")
gs.cell(row=G_HEAD_ROW, column=3, value="状態")
gs.cell(row=G_HEAD_ROW, column=4, value="販売開始")
for w in range(WEEKS):
    col = 5 + w
    gs.cell(row=G_HEAD_ROW, column=col,
            value="=$B$2" if w == 0 else f"={L(col-1)}{G_HEAD_ROW}+7")
    gs.cell(row=G_HEAD_ROW, column=col).number_format = "m/d"
style_header(gs, G_HEAD_ROW, 4 + WEEKS, height=26)
widths(gs, {"A": 26, "B": 8, "C": 10, "D": 12})
for w in range(WEEKS):
    gs.column_dimensions[L(5 + w)].width = 3.4
gs.freeze_panes = "E4"

for i in range(2, MAX_S + 1):
    r = G_HEAD_ROW + i - 1              # 学校マスタ row i ↔ ガント row r
    gs.cell(row=r, column=1, value=f'=IF(学校マスタ!$A{i}="","",学校マスタ!$A{i})').font = Font(name=FONT, size=10, color=GREEN)
    gs.cell(row=r, column=2, value=f'=IF(学校マスタ!$A{i}="","",学校マスタ!$C{i})').font = Font(name=FONT, size=10, color=GREEN)
    gs.cell(row=r, column=3, value=f'=IF(学校マスタ!$A{i}="","",学校マスタ!$X{i})').font = Font(name=FONT, size=10, color=GREEN)
    c = gs.cell(row=r, column=4, value=f'=IF(学校マスタ!$A{i}="","",学校マスタ!$R{i})')
    c.font = Font(name=FONT, size=10, color=GREEN); c.number_format = "m/d"
    for col in range(1, 5):
        gs.cell(row=r, column=col).border = box
    for w in range(WEEKS):
        col = 5 + w
        cl = L(col)
        milestone = ",".join([
            f'AND(学校マスタ!${DATE[k]}{i}<>"",学校マスタ!${DATE[k]}{i}>={cl}${G_HEAD_ROW},学校マスタ!${DATE[k]}{i}<{cl}${G_HEAD_ROW}+7)'
            for k in ("model", "intent", "ec")])
        f = (f'=IF(学校マスタ!$A{i}="","",'
             f'IF(AND(学校マスタ!$U{i}<>"",学校マスタ!$U{i}>={cl}${G_HEAD_ROW},学校マスタ!$U{i}<{cl}${G_HEAD_ROW}+7),"◎",'
             f'IF(OR({milestone}),"◆",'
             f'IF(AND(学校マスタ!$O{i}<>"",学校マスタ!$P{i}<>"",学校マスタ!$O{i}<{cl}${G_HEAD_ROW}+7,学校マスタ!$P{i}>={cl}${G_HEAD_ROW}),"█",'
             f'IF(AND(学校マスタ!$R{i}<>"",学校マスタ!$U{i}<>"",学校マスタ!$R{i}<{cl}${G_HEAD_ROW}+7,学校マスタ!$U{i}>={cl}${G_HEAD_ROW}),"▒","")))))')
        cell = gs.cell(row=r, column=col, value=f)
        cell.font = Font(name=FONT, size=9, color=BLACK)
        cell.alignment = Alignment(horizontal="center")
last_g = G_HEAD_ROW + MAX_S - 1
grng = f"E{G_HEAD_ROW + 1}:{L(4 + WEEKS)}{last_g}"
for sym, color, fill in [("█", "2F4B8F", "DCE3F5"), ("▒", "0F5E58", "D6EDEA"),
                         ("◆", "8A5A12", "FBEED2"), ("◎", "9B2C1E", "F8D7D5")]:
    gs.conditional_formatting.add(grng, CellIsRule(operator="equal", formula=[f'"{sym}"'],
        font=Font(name=FONT, size=9, bold=True, color=color), fill=PatternFill("solid", fgColor=fill)))
gs.conditional_formatting.add(f"E{G_HEAD_ROW}:{L(4 + WEEKS)}{G_HEAD_ROW}", FormulaRule(
    formula=[f'AND(E${G_HEAD_ROW}<=TODAY(),TODAY()<E${G_HEAD_ROW}+7)'],
    fill=PatternFill("solid", fgColor="B33A2B"), font=Font(name=FONT, size=10, bold=True, color="FFFFFF")))
gs.conditional_formatting.add(grng, FormulaRule(
    formula=[f'AND(E${G_HEAD_ROW}<=TODAY(),TODAY()<E${G_HEAD_ROW}+7,E{G_HEAD_ROW + 1}="")'],
    fill=PatternFill("solid", fgColor="F6E4E1")))

# ---------------------------------------------------------------- 手数料集計
fs = wb.create_sheet("手数料集計")
fs.sheet_view.showGridLines = False
fs["A1"] = "全体サマリー"
fs["A1"].font = Font(name=FONT, bold=True, size=13, color=HEAD)
summary = [
    ("対象校", f'=COUNTA(学校マスタ!$A$2:$A${MAX_S})', "0"),
    ("遅延あり", f'=COUNTIF(学校マスタ!$X$2:$X${MAX_S},"遅延")', "0"),
    ("今週対応", f'=COUNTIF(学校マスタ!$X$2:$X${MAX_S},"今週対応")', "0"),
    ("価格 未入力の学校", f'=COUNTIF(学校マスタ!$AA$2:$AA${MAX_S},">0")', "0"),
    ("手数料バック なしの学校", f'=COUNTIF(学校マスタ!$G$2:$G${MAX_S},"なし")', "0"),
    ("手数料バック 見込 合計", f'=SUM(価格!$N$2:$N${MAX_P})', '¥#,##0'),
]
for i, (label, f, fmt) in enumerate(summary, start=2):
    a = fs.cell(row=i, column=1, value=label)
    a.font = Font(name=FONT, size=10, color=INK); a.border = box
    b = fs.cell(row=i, column=2, value=f)
    b.font = Font(name=FONT, bold=True, size=11, color=GREEN); b.border = box
    b.number_format = fmt
    b.alignment = Alignment(horizontal="right")
widths(fs, {"A": 26, "B": 18, "C": 12, "D": 16, "E": 16, "F": 16})

fs["A9"] = "協力企業に戻す手数料（見込み・税別）"
fs["A9"].font = Font(name=FONT, bold=True, size=13, color=HEAD)
head = ["協力企業", "対象校", "PC", "オプション", "合計"]
for i, hcell in enumerate(head, start=1):
    fs.cell(row=10, column=i, value=hcell)
style_header(fs, 10, len(head), height=22)
FIRST, LAST = 11, 20
for i, r in enumerate(range(FIRST, LAST + 1), start=2):
    fs.cell(row=r, column=1, value=f'=IF(設定!$C{i}="","",設定!$C{i})').font = Font(name=FONT, size=10, color=GREEN)
    fs.cell(row=r, column=2, value=f'=IF($A{r}="","",COUNTIFS(学校マスタ!$F$2:$F${MAX_S},$A{r},学校マスタ!$G$2:$G${MAX_S},"あり"))')
    fs.cell(row=r, column=3, value=f'=IF($A{r}="","",SUMIFS(価格!$N$2:$N${MAX_P},価格!$P$2:$P${MAX_P},$A{r},価格!$B$2:$B${MAX_P},"PC"))')
    fs.cell(row=r, column=4, value=f'=IF($A{r}="","",SUMIFS(価格!$N$2:$N${MAX_P},価格!$P$2:$P${MAX_P},$A{r},価格!$B$2:$B${MAX_P},"オプション"))')
    fs.cell(row=r, column=5, value=f'=IF($A{r}="","",SUM($C{r}:$D{r}))')
    for col in range(1, 6):
        c = fs.cell(row=r, column=col)
        c.border = box
        if col >= 2:
            c.font = Font(name=FONT, size=10, color=BLACK)
            c.number_format = '¥#,##0;-¥#,##0;"-"' if col >= 3 else '0;;"-"'
            c.alignment = Alignment(horizontal="right")
tr = LAST + 1
fs.cell(row=tr, column=1, value="合計").font = Font(name=FONT, bold=True, size=10, color=INK)
for col in range(2, 6):
    c = fs.cell(row=tr, column=col, value=f'=SUM({L(col)}{FIRST}:{L(col)}{LAST})')
    c.font = Font(name=FONT, bold=True, size=10, color=BLACK)
    c.number_format = '¥#,##0;-¥#,##0;"-"' if col >= 3 else '0;;"-"'
    c.alignment = Alignment(horizontal="right")
for col in range(1, 6):
    fs.cell(row=tr, column=col).fill = PatternFill("solid", fgColor=BAND)
    fs.cell(row=tr, column=col).border = box
fs.cell(row=tr + 2, column=1,
        value="見込合計 ＝（今年 掲載（エンド税別）− 今年 卸（税別））× 見込 数量。手数料バックが「なし」の学校は 0 として扱われます。"
        ).font = Font(name=FONT, size=9, color=MUTED)
fs.cell(row=tr + 3, column=1,
        value="協力企業は「設定」シートのC列から取っています。増やすときは設定シートに追記してください。"
        ).font = Font(name=FONT, size=9, color=MUTED)

# ---------------------------------------------------------------- サンプル1校
sample = ["青葉学園高等学校", "東京", "内野", "継続", 320, "サクラ電機販売", "あり",
          dt.date(2026,7,8), "✓", dt.date(2026,7,25), "✓", dt.date(2026,8,4), "", "",
          dt.date(2026,8,1), dt.date(2026,8,11), "", dt.date(2026,8,14), "",
          "業者直送（分割）", dt.date(2026,9,1), "✓"]
for i, v in enumerate(sample, start=1):
    if v != "":
        ms.cell(row=2, column=i, value=v)
ms.cell(row=2, column=29, value="サンプル行です。実データを入れる前に削除してください。")

items = [
    ("PC", "Bモデル ノート13型 軽量 / i5・16GB", 169000, 135000, 155, 172000, 140000, 160),
    ("PC", "Aモデル ノート14型 / i5・16GB", None, None, None, None, None, 120),
    ("オプション", "キャリングケース", 4800, 3200, None, 4800, 3200, 101),
    ("オプション", "Office 永続ライセンス", 32000, None, None, 32000, 24000, 121),
]
for j, (kind, name, dp, wp, up, hn, jn, mp) in enumerate(items, start=2):
    ps.cell(row=j, column=1, value="青葉学園高等学校")
    ps.cell(row=j, column=2, value=kind)
    ps.cell(row=j, column=3, value=name)
    for col, val in ((4, dp), (5, wp), (6, up), (8, hn), (10, jn), (13, mp)):
        if val is not None:
            ps.cell(row=j, column=col, value=val)

wb.save(OUT)
print("saved", OUT)
